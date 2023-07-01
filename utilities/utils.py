import inspect
import softest
import logging
from openpyxl import Workbook, load_workbook
import csv


#anything independent from the driver method goes to the utility
class Utils(softest.TestCase):
    def assertListItemText(self, list, value):
        for stop in list:
            print("The text is: " + stop.text)
            self.soft_assert(self.assertEqual, stop.text, value) #this value can be the 1stop or 2 stop
            if stop.text == value:
                print("test passed")
            else:
                 print("test failed")
        self.assert_all() #it works like a charm


        def custom_logger(logLevel=logging.DEBUG):
            #set class/method name from where its called
            logger_name = inspect.stack()[1][3]
            # create logger
            logger = logging.getLogger(logger_name)  # it's best to use name because it will automatically pick up the name from the import folder that is; where the file is coming from and we add LoggerDemo to display where the particular classes are coming from
            logger.setLevel(logLevel)

            # create console handler or file handler and set the log level
            fh = logging.FileHandler("automation.log", mode='a') # normal mode is 'w' thus it will over write the files. u can create different files for diff class but it good to keep all the log into one class in automation

            # create formatter - how you want your logs to be formatted
            formatter = logging.Formatter('%(asctime)s - %(levelname)s : %(message)s', datefmt='%m/%d/%Y %I:%M:%S %p')

            # add formatter to console or file handler
            fh.setFormatter(formatter)

            # add console handler to logger
            logger.addHandler(fh)
            return logger

        def read_data_from_excel(file_name, sheet):
            datalist = []
            wb = load_workbook(filename=file_name)
            sh = wb[sheet]
            row_ct = sh.max_row
            col_ct = sh.max_column

            for i in range(2, row_ct + 1): #be aware we don't put the header number in the max row
                row = []
                for j in range(1, col_ct + 1):
                    row.append(sh.cell(row=i, column=j).value)
                datalist.append(row)
            return datalist

        def read_data_from_csv(filename):
            #create an empty list
            datalist = []

            #open CSV file
            csvdata = open(filename,"r")

            #create csv reader
            reader = csv.reader(csvdata)

            #skip header
            next(reader)

            #add CSV rows to list
            for rows in reader:
                datalist.append(rows)
            return datalist




            #exception handeling should be implemented if need be


            # application code - log messages
            #logger.debug("debug log statement")
            #logger.info("info log statement")
            #logger.warning("warning log statement")
            #logger.error("error log statement")
            #logger.critical("critical log statement"












